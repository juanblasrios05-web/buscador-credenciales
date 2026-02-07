#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para generar datos.json correctamente desde planilla.xlsx
Universidad UMAX - Buscador de Credenciales
"""

import openpyxl
import json
import re

print("=" * 60)
print("GENERADOR DE JSON - BUSCADOR CREDENCIALES UMAX")
print("=" * 60)

# Cargar Excel
print("\n📂 Cargando planilla.xlsx...")
wb = openpyxl.load_workbook('planilla.xlsx', data_only=True)
ws = wb.active

datos = []
errores = []

print(f"✓ Archivo cargado. Procesando {ws.max_row - 1} registros...\n")

# Mapeo correcto de columnas
# A: Nº
# B: COD ALUM (código del estudiante)
# C: USUARIO BIBLIOTECA
# D: CONTRASEÑA BIBLIO
# E: NOMBRE
# F: APELLIDO
# G: GRUPO
# H: USUARIO AV ← ESTE SE USA PARA BUSCAR
# I: USUARIO PORTAL
# J: CORREO
# K: CONTRASEÑA2

for row_num in range(2, ws.max_row + 1):
    try:
        # Obtener valores de la fila
        row = ws[row_num]
        
        # Extraer valores (índice 0-based)
        num = row[0].value  # Columna A: Nº
        codigo_estudiante = row[1].value  # Columna B: COD ALUM
        usuario_biblio_excel = row[2].value  # Columna C: USUARIO BIBLIOTECA
        password_biblio = row[3].value  # Columna D: CONTRASEÑA BIBLIO
        nombre = row[4].value  # Columna E: NOMBRE
        apellido = row[5].value  # Columna F: APELLIDO
        grupo = row[6].value  # Columna G: GRUPO
        usuario_av_excel = row[7].value  # Columna H: USUARIO AV
        usuario_portal_excel = row[8].value  # Columna I: USUARIO PORTAL
        correo_excel = row[9].value  # Columna J: CORREO ELECTRÓNICO
        password_general = row[10].value  # Columna K: CONTRASEÑA
        
        # Validar que tenga al menos código y nombre
        if not codigo_estudiante or not nombre:
            errores.append(f"Fila {row_num}: Faltan datos críticos (código o nombre)")
            continue
        
        # Limpiar y procesar datos
        codigo_clean = str(codigo_estudiante).strip() if codigo_estudiante else ""
        nombre_clean = str(nombre).strip().upper() if nombre else ""
        apellido_clean = str(apellido).strip().upper() if apellido else ""
        grupo_clean = str(grupo).strip().upper() if grupo else ""
        
        # ========================================
        # PROCESAMIENTO DE USUARIOS Y CONTRASEÑAS
        # ========================================
        
        # 1. USUARIO BIBLIOTECA (Columna C)
        usuario_biblio_clean = str(usuario_biblio_excel).strip() if usuario_biblio_excel else ""
        
        # 2. CONTRASEÑA BIBLIOTECA (Columna D)
        pass_biblio_clean = str(password_biblio).strip() if password_biblio else "Alumno.2021"
        
        # 3. USUARIO AULA VIRTUAL (Columna H) - ESTE ES EL QUE USARÁN PARA BUSCAR
        usuario_av_clean = str(usuario_av_excel).strip() if usuario_av_excel else usuario_biblio_clean
        
        # 4. USUARIO PORTAL (Columna I)
        usuario_portal_clean = str(usuario_portal_excel).strip() if usuario_portal_excel else usuario_biblio_clean
        
        # 5. CORREO ELECTRÓNICO (Columna J)
        if correo_excel and str(correo_excel).strip():
            correo_clean = str(correo_excel).strip()
        else:
            # Generar correo si no existe
            if nombre_clean and apellido_clean:
                primer_nombre = nombre_clean.split()[0].lower()
                primer_apellido = apellido_clean.split()[0].lower()
                correo_clean = f"{primer_nombre}.{primer_apellido}@umax.edu.py"
                # Quitar acentos
                correo_clean = correo_clean.replace('á', 'a').replace('é', 'e').replace('í', 'i')
                correo_clean = correo_clean.replace('ó', 'o').replace('ú', 'u').replace('ñ', 'n')
            else:
                correo_clean = ""
        
        # 6. CONTRASEÑA GENERAL (Columna K) - Para AV, Portal y Correo
        pass_general_clean = str(password_general).strip() if password_general else "Alumno.2021"
        if pass_general_clean in ["None", "nan", ""]:
            pass_general_clean = "Alumno.2021"
        
        # Crear registro JSON con todos los datos
        registro = {
            "codigo": codigo_clean,                    # Columna B: COD ALUM
            "nombre": nombre_clean,                    # Columna E
            "apellido": apellido_clean,                # Columna F
            "grupo": grupo_clean,                      # Columna G
            "usuario_biblio": usuario_biblio_clean,    # Columna C
            "pass_biblio": pass_biblio_clean,          # Columna D
            "usuario_av": usuario_av_clean,            # Columna H ← ESTE ES PARA BUSCAR
            "usuario_portal": usuario_portal_clean,    # Columna I
            "correo": correo_clean,                    # Columna J
            "pass_general": pass_general_clean         # Columna K
        }
        
        datos.append(registro)
        
        # Mostrar progreso cada 100 registros
        if len(datos) % 100 == 0:
            print(f"  ⏳ Procesados {len(datos)} registros...")
        
    except Exception as e:
        errores.append(f"Fila {row_num}: {str(e)}")
        continue

# Resumen
print(f"\n{'=' * 60}")
print(f"✅ PROCESAMIENTO COMPLETADO")
print(f"{'=' * 60}")
print(f"📊 Total registros procesados: {len(datos)}")
print(f"⚠️  Errores encontrados: {len(errores)}")

if errores:
    print(f"\n⚠️  ERRORES DETECTADOS:")
    for error in errores[:10]:
        print(f"   - {error}")
    if len(errores) > 10:
        print(f"   ... y {len(errores) - 10} errores más")

# Mostrar ejemplos
if datos:
    print(f"\n{'=' * 60}")
    print("📋 EJEMPLO DE REGISTROS GENERADOS:")
    print(f"{'=' * 60}")
    
    for i in range(min(3, len(datos))):
        print(f"\n{'-' * 60}")
        print(f"Registro {i+1}:")
        print(json.dumps(datos[i], indent=2, ensure_ascii=False))

# Guardar JSON
output_file = 'datos_nuevo.json'
print(f"\n{'=' * 60}")
print(f"💾 Guardando archivo: {output_file}")
print(f"{'=' * 60}")

with open(output_file, 'w', encoding='utf-8') as f:
    json.dump(datos, f, ensure_ascii=False, indent=2)

print(f"✅ Archivo guardado exitosamente")
print(f"📦 Tamaño: {len(datos)} registros")
print(f"\n{'=' * 60}")
print("🎉 PROCESO COMPLETADO")
print(f"{'=' * 60}\n")

# Validaciones adicionales
print("🔍 VALIDACIONES:")
print(f"   ✓ Registros con código válido: {sum(1 for d in datos if d['codigo'])}")
print(f"   ✓ Registros con nombre completo: {sum(1 for d in datos if d['nombre'] and d['apellido'])}")
print(f"   ✓ Registros con correo: {sum(1 for d in datos if d['correo'])}")
print(f"   ✓ Registros con usuario AV: {sum(1 for d in datos if d['usuario_av'])}")
print(f"\n")