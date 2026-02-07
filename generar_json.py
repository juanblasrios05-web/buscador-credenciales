#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para generar datos.json correctamente desde planilla.xlsx
Universidad UMAX - Buscador de Credenciales
"""

import openpyxl
import json
import re

print("=" * 60)
print("GENERADOR DE JSON - BUSCADOR CREDENCIALES UMAX")
print("=" * 60)

# Cargar Excel
print("\n📂 Cargando planilla.xlsx...")
wb = openpyxl.load_workbook('/mnt/user-data/uploads/planilla.xlsx', data_only=True)
ws = wb.active

datos = []
errores = []

print(f"✓ Archivo cargado. Procesando {ws.max_row - 1} registros...\n")

# Mapeo correcto de columnas (según análisis del Excel real)
# A: Nº (índice)
# B: COD ALUM (pero está vacío, usar C)
# C: CODIGO REAL (10114113, 5959397, etc.)
# D: USUARIO BIBLIOTECA (umax@113, umax@397, etc.)
# E: NOMBRE
# F: APELLIDO
# G: GRUPO
# H: USUARIO AV (vacío)
# I: USUARIO PORTAL (vacío)
# J: CORREO (vacío)
# K: CONTRASEÑA2

for row_num in range(2, ws.max_row + 1):
    try:
        # Obtener valores de la fila
        row = ws[row_num]
        
        # Extraer valores (índice 0-based)
        num = row[0].value  # Columna A: Nº
        cod_alum_b = row[1].value  # Columna B: COD ALUM (vacío)
        codigo_real = row[2].value  # Columna C: CODIGO REAL
        usuario_biblio_raw = row[3].value  # Columna D: USUARIO BIBLIOTECA (umax@113)
        nombre = row[4].value  # Columna E: NOMBRE
        apellido = row[5].value  # Columna F: APELLIDO
        grupo = row[6].value  # Columna G: GRUPO
        usuario_av = row[7].value  # Columna H: USUARIO AV (vacío)
        usuario_portal = row[8].value  # Columna I: USUARIO PORTAL (vacío)
        correo = row[9].value  # Columna J: CORREO (vacío)
        password = row[10].value  # Columna K: CONTRASEÑA2
        
        # Validar que tenga al menos código y nombre
        if not codigo_real or not nombre:
            errores.append(f"Fila {row_num}: Faltan datos críticos (código o nombre)")
            continue
        
        # Limpiar y procesar datos
        codigo_clean = str(codigo_real).strip() if codigo_real else ""
        codigo_clean = re.sub(r'[^\d]', '', codigo_clean)  # Solo números
        
        nombre_clean = str(nombre).strip().upper() if nombre else ""
        apellido_clean = str(apellido).strip().upper() if apellido else ""
        grupo_clean = str(grupo).strip().upper() if grupo else ""
        
        # Procesar usuario biblioteca
        # Formato original: "umax@113" -> extraer "113"
        usuario_biblio_num = ""
        if usuario_biblio_raw:
            match = re.search(r'@?(-?\d+)', str(usuario_biblio_raw))
            if match:
                usuario_biblio_num = match.group(1)
        
        # Si no hay usuario_biblio, extraer últimos 3 dígitos del código
        if not usuario_biblio_num and codigo_clean:
            usuario_biblio_num = codigo_clean[-3:] if len(codigo_clean) >= 3 else codigo_clean
        
        usuario_biblio = f"umax{usuario_biblio_num}"
        usuario_av_clean = usuario_biblio_num
        
        # Generar correo: nombre.apellido@umax.edu.py
        if nombre_clean and apellido_clean:
            # Tomar primer nombre y primer apellido
            primer_nombre = nombre_clean.split()[0].lower()
            primer_apellido = apellido_clean.split()[0].lower()
            
            # Quitar acentos para el correo
            correo_generado = f"{primer_nombre}.{primer_apellido}@umax.edu.py"
            correo_generado = correo_generado.replace('á', 'a').replace('é', 'e').replace('í', 'i')
            correo_generado = correo_generado.replace('ó', 'o').replace('ú', 'u').replace('ñ', 'n')
        else:
            correo_generado = ""
        
        # Usuario portal es el código limpio
        usuario_portal_clean = codigo_clean
        
        # Contraseña (por defecto Alumno.2021)
        password_clean = str(password).strip() if password else "Alumno.2021"
        if password_clean in ["None", "nan", ""]:
            password_clean = "Alumno.2021"
        
        # Crear registro
        registro = {
            "codigo": codigo_clean,
            "nombre": nombre_clean,
            "apellido": apellido_clean,
            "grupo": grupo_clean,
            "usuario_biblio": usuario_biblio,
            "pass_biblio": "Alumno.2021",
            "usuario_av": usuario_av_clean,
            "usuario_portal": usuario_portal_clean,
            "correo": correo_generado,
            "pass_general": password_clean
        }
        
        datos.append(registro)
        
        # Mostrar progreso cada 100 registros
        if len(datos) % 100 == 0:
            print(f"  ⏳ Procesados {len(datos)} registros...")
        
    except Exception as e:
        errores.append(f"Fila {row_num}: {str(e)}")
        continue

# Resumen
print(f"\n{'=' * 60}")
print(f"✅ PROCESAMIENTO COMPLETADO")
print(f"{'=' * 60}")
print(f"📊 Total registros procesados: {len(datos)}")
print(f"⚠️  Errores encontrados: {len(errores)}")

if errores:
    print(f"\n⚠️  ERRORES DETECTADOS:")
    for error in errores[:10]:  # Mostrar solo primeros 10
        print(f"   - {error}")
    if len(errores) > 10:
        print(f"   ... y {len(errores) - 10} errores más")

# Mostrar ejemplos
if datos:
    print(f"\n{'=' * 60}")
    print("📋 EJEMPLO DE REGISTROS GENERADOS:")
    print(f"{'=' * 60}")
    
    for i in range(min(3, len(datos))):
        print(f"\n{'-' * 60}")
        print(f"Registro {i+1}:")
        print(json.dumps(datos[i], indent=2, ensure_ascii=False))

# Guardar JSON
output_file = '/home/claude/datos_nuevo.json'
print(f"\n{'=' * 60}")
print(f"💾 Guardando archivo: {output_file}")
print(f"{'=' * 60}")

with open(output_file, 'w', encoding='utf-8') as f:
    json.dump(datos, f, ensure_ascii=False, indent=2)

print(f"✅ Archivo guardado exitosamente")
print(f"📦 Tamaño: {len(datos)} registros")
print(f"\n{'=' * 60}")
print("🎉 PROCESO COMPLETADO")
print(f"{'=' * 60}\n")

# Validaciones adicionales
print("🔍 VALIDACIONES:")
print(f"   ✓ Registros con código válido: {sum(1 for d in datos if d['codigo'])}")
print(f"   ✓ Registros con nombre completo: {sum(1 for d in datos if d['nombre'] and d['apellido'])}")
print(f"   ✓ Registros con correo generado: {sum(1 for d in datos if d['correo'])}")
print(f"   ✓ Registros con usuario biblioteca: {sum(1 for d in datos if d['usuario_biblio'])}")
print(f"\n")
